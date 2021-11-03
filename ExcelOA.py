import openpyxl
import matplotlib.pyplot as plt
import xlrd
from pyecharts.charts import Bar,Line,Pie,Scatter
from pyecharts import options as opts

import numpy as np
from openpyxl.chart import (
    Reference,
    series,
    PieChart,
    BarChart,
    BubbleChart
)
import xlwt

# wb = openpyxl.load_workbook(r'E:\\pytextOA\\people.xlsx')
# # print(wb.sheetnames)
# ws = wb.active
# ws.title = 'pieChart'
# data=[['Pie','Sold'],
#       ['韩国',120750],
#       ['美国',71493],
#       ['日本',66159],
#       ['缅甸',39776],
#       ['越南',36205]
#       ]
# for row in data:
#     ws.append(row)
# pie =PieChart()
# labels = Reference(ws,min_col=1,min_row=2,max_row=5)
# data = Reference(ws,min_col=2,min_row=2,max_row=5)
# pie.add_data(data)
# pie.set_categories(labels)
# pie.title='test map'
# ws.add_chart(pie,'A20')
# print(ws['A1'].value)
# 打开文件
# wb = xlrd.open_workbook(r'example.xlsx')
# print(wb.sheetnames)
# mySheet=wb.create_sheet('MySheet')
# 创建一个表单
# for sheet in wb:
#     print(sheet.title)
# # 循环打出表单名字
# sheet3 = wb.get_sheet_by_name('Sheet1')
# sheet4 = wb['MySheet']
# # 分别以两种方式获取表单对象
# ws = wb.active
# print(ws)
# print(ws['A'])
# print(ws['A'.value])
# xlsx = xlrd.open_workbook('E:\\pytextOA\\example1.xls')
# table = xlsx.sheet_by_name('第一个工作簿')
# print(table.cell_value(0,3))
# print(table.cell(1,2).value)
# print(table.row(1)[2].value)
# #/////////////////////////////////////////////
# #写操作
# new_Workbook = xlwt.Workbook()
# worksheet =new_Workbook.add_sheet('new_test')
# worksheet.write(0,0,'test')
# new_Workbook.save('E:\\pytextOA\\example1.xls')
# ws = wb.active
# print(ws['A1'].value)
# print(ws.cell(row=1,column=1).value)
# for row in ws.iter_rows(min_row=1,max_row=2,max_rol=1):
#     for cell in row:
#         print(cell)
# data_list = []
# with xlrd.open_workbook(r'E:\\pytextOA\\people.xlsx') as data:
#     table =data.sheet_by_index(0)
#     # 取得有多少行数据
#     rows_count =table.nrows
#     # 第一行不要 是表头
#     for row_idx in range(1,rows_count):
#         # 序号为1的 第二个
#         contry = table.cell(row_idx,0).value
#         unit_saled = table.cell(row_idx,1).value
#         # print(f"{contry}:{unit_saled}",end=',')
#         data_list.append((contry,unit_saled))
# # # 去重
# # contrys = list(set([contry for contry,ct in data_list]))
# # # 人数相加
# # result =[(contry,sum([s for c,s in data_list if contry==c ]))]
# # for contry,rs in result:
# #     print(f'{contry:9}:{rs:9}')
# labels = contry
# sizes = unit_saled
# plt.pie(sizes,labels=labels,)
# plt.axis('equal')
# plt.show()
with xlrd.open_workbook(r'E:\\pytextOA\\people.xlsx') as data:
    table = data.sheets()[0]
    y = []
    x = []
    y = table.col_values(1)
    x = table.col_values(0)
    print(y)
    print(x)
    bar =(
        Bar()
        .add_xaxis(x)
        .add_yaxis("人数",y)
        # .add_yaxis("人数", [5,20,30,10,75])
        .set_global_opts(title_opts=opts.TitleOpts(title="主标题",subtitle="副标题"))
    )
    bar.render('picture.html')
  # t = np.linspace(1, 296, len(y))
    # bar = Bar("各国外籍人口展示","统计如下")
    # bar.add("折线图展示",t,y,is_more_utils=True)
    # bar.show_config()
    # bar.render(r"E:\\zhanshi\\bokezhexiantu.html")